import csv
import os
from pypdf import PdfReader
from openpyxl import load_workbook


def test_laborers_csv(create_new_archive):
    path = os.path.join(create_new_archive, 'testОснова.csv')
    with open(path, encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=";")
        assert [row['Должность'] for row in reader] == ['Продавец']



def test_users_xlsx(create_new_archive):
    path = os.path.join(create_new_archive, 'import_empl_xlsx.xlsx')
    open_xlsx = load_workbook(path)
    sheet = open_xlsx.active
    name = sheet.cell(row=3, column=2).value
    assert name == 'Сергеев'




def test_book_pdf(create_new_archive):
    path = f'{create_new_archive}/AByteofPythonRussian-2.02.pdf'
    text_to_search = "This book about PHPfghdfg"
    with open(path, 'rb') as file:
        reader = PdfReader(file)
        text_found = any(text_to_search in page.extract_text() for page in reader.pages)
        assert text_to_search, f"Текст \"{text_found}\" не найден в PDF файле."

